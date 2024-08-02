#registros/views.py

import pandas as pd
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from .forms import RegistroHorasExtrasForm
import os
import tempfile
import logging
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, PageBreak
from io import BytesIO

logger = logging.getLogger(__name__)

def autocomplete_funcionarios(request):
    if request.method == 'GET':
        term = request.GET.get('term', '')
        excel_path = os.path.join(os.path.dirname(__file__), r'D:\1Desktop\Documentos\My Web Sites\App py\ProjetoCastan\Funcionários.xlsx')
        
        funcionarios_df = pd.read_excel(excel_path)
        nomes = funcionarios_df[funcionarios_df['Nome do funcionário'].str.startswith(term, na=False)]['Nome do funcionário'].tolist()
        
        return JsonResponse(nomes, safe=False)
    
def formatar_numero(numero):
    """
    Formata um número como string com máscara de milhar e decimal (1.000,00).
    """
    if isinstance(numero, (int, float)):
        return f"{numero:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    return numero

    
def desenhar_tabela(pdf_canvas, dados):
    data = [['Atributo', 'Valor']]
    data += [[chave, f"{valor:.2f}".replace('.', ',')] if isinstance(valor, (int, float)) else [chave, valor] for chave, valor in dados.items()]

    table = Table(data, colWidths=[2.5 * inch, 2.5 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements = []
    elements.append(table)
    
    pdf_canvas.build(elements)

@csrf_exempt
def registro_horas_extras(request):
    if request.method == 'POST':
        form = RegistroHorasExtrasForm(request.POST)
        if form.is_valid():
            nome_funcionario = form.cleaned_data['nome_funcionario']
            excel_path = os.path.join(os.path.dirname(__file__), r'D:\1Desktop\Documentos\My Web Sites\App py\ProjetoCastan\Funcionários.xlsx')
            funcionarios_df = pd.read_excel(excel_path)
            pessoa_selecionada = funcionarios_df[funcionarios_df['Nome do funcionário'] == nome_funcionario]

            if not pessoa_selecionada.empty:
                salario = pessoa_selecionada.iloc[0]['Salário']
                carga_horaria = pessoa_selecionada.iloc[0]['Carga Horária']

                try:
                    salario = float(salario)
                    carga_horaria = float(carga_horaria)
                except ValueError:
                    return JsonResponse({'error': 'Salário ou carga horária inválidos.'}, status=400)
                
                salario_por_hora = salario / carga_horaria
                dias_uteis = int(form.cleaned_data["dias_uteis"])
                dsr = int(form.cleaned_data["dsr"])

                he60_qtde = float(form.cleaned_data["he60_qtde"])
                he80_qtde = float(form.cleaned_data["he80_qtde"])
                he80_qtde_noturno = float(form.cleaned_data["he80_qtde_noturno"])
                he100_qtde = float(form.cleaned_data["he100_qtde"])

                he60_valor = salario_por_hora * 1.6 * he60_qtde
                he80_valor = salario_por_hora * 1.8 * he80_qtde
                he80_valor_noturno = (salario_por_hora * 1.8 * 1.3) * (he80_qtde_noturno * 1.1428571)
                he100_valor = salario_por_hora * 2 * he100_qtde

                refl_dsr = ((he60_valor + he80_valor + he80_valor_noturno + he100_valor) / dias_uteis) * dsr
                total_he = he60_valor + he80_valor + he80_valor_noturno + he100_valor + refl_dsr

                decimo_terceiro = (salario_por_hora * total_he) / 12
                ferias = (salario_por_hora * total_he) / 12
                um_terco_ferias = ferias / 3
                fgts = (total_he + decimo_terceiro + ferias + um_terco_ferias) * 0.08

                ferias_faturado = total_he / 12 * 1.333333
                decimo_terceiro_faturado = total_he / 12
                inss = (total_he + ferias_faturado + decimo_terceiro_faturado) * 0.28
                fgts_faturado = (total_he + ferias_faturado + decimo_terceiro_faturado) * 0.08
                total_faturado = total_he + ferias_faturado + decimo_terceiro_faturado + inss + fgts_faturado
                fgst_40 = fgts_faturado * 0.4
                total_absoluto = total_faturado + fgst_40
                valor_nota = total_absoluto / 0.8

                dados = {
                    'Nome do Funcionário': nome_funcionario,
                    'Salário': formatar_numero(salario),
                    'Carga Horária': formatar_numero(carga_horaria),
                    'Salário por Hora': formatar_numero(salario_por_hora),
                    'Décimo Terceiro': formatar_numero(decimo_terceiro),
                    'Férias': formatar_numero(ferias),
                    '1/3 de Férias': formatar_numero(um_terco_ferias),
                    'FGTS': formatar_numero(fgts),
                    'Reflexo DSR': formatar_numero(refl_dsr),
                    'Total de Horas Extras': formatar_numero(total_he),
                    'Férias Faturado': formatar_numero(ferias_faturado),
                    'Décimo Terceiro Faturado': formatar_numero(decimo_terceiro_faturado),
                    'INSS': formatar_numero(inss),
                    'FGTS Faturado': formatar_numero(fgts_faturado),
                    'Total Faturado': formatar_numero(total_faturado),
                    'FGTS 40%': formatar_numero(fgst_40),
                    'Total Absoluto': formatar_numero(total_absoluto),
                    'Valor da Nota': formatar_numero(valor_nota)
                }

                if 'download_pdf' in request.POST:
                    buffer = BytesIO()
                    doc = SimpleDocTemplate(buffer, pagesize=letter)
                    elements = []

                    data = [['Atributo', 'Valor']]
                    data += [[chave, formatar_numero(valor)] if isinstance(valor, (int, float)) else [chave, valor] for chave, valor in dados.items()]

                    table = Table(data, colWidths=[2.5 * inch, 2.5 * inch])
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ]))

                    elements.append(table)
                    doc.build(elements)

                    buffer.seek(0)
                    response = HttpResponse(buffer, content_type='application/pdf')
                    response['Content-Disposition'] = 'attachment; filename="horas_extras.pdf"'
                    return response

                if 'download_excel' in request.POST:
                    df_resultados = pd.DataFrame(list(dados.items()), columns=['Atributo', 'Valor'])
                    
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                        df_resultados.to_excel(temp_file.name, index=False)
                        temp_file_path = temp_file.name

                    # Corrige a formatação no arquivo Excel
                    workbook = load_workbook(temp_file_path)
                    sheet = workbook.active

                    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
                        for cell in row:
                            if isinstance(cell.value, (int, float)):
                                cell.value = formatar_numero(cell.value).replace('.', ',')  # Substitui ponto por vírgula

                    workbook.save(temp_file_path)

                    with open(temp_file_path, 'rb') as excel_file:
                        response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        response['Content-Disposition'] = 'attachment; filename="resultados_calculos.xlsx"'

                    os.remove(temp_file_path)
                    return response

                return JsonResponse(dados)
            else:
                return JsonResponse({'error': 'Funcionário não encontrado.'}, status=404)
        else:
            return JsonResponse({'error': 'Formulário inválido.'}, status=400)
    else:
        excel_path = os.path.join(os.path.dirname(__file__), r'D:\1Desktop\Documentos\My Web Sites\App py\ProjetoCastan\Funcionários.xlsx')
        funcionarios_df = pd.read_excel(excel_path)
        nomes_funcionarios = funcionarios_df['Nome do funcionário'].tolist()
        
        choices = [(nome, nome) for nome in nomes_funcionarios]
        
        form = RegistroHorasExtrasForm()
        form.fields['nome_funcionario'].choices = choices
    
    return render(request, 'registro_horas_extras.html', {'form': form, 'names': nomes_funcionarios})

def sucesso(request):
    return render(request, 'sucesso.html')

def calculados(request):
    context = {
        # Dados que você deseja passar para o template
    }
    return render(request, 'calculados.html', context)
